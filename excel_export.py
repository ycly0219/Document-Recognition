"""Excel 识别结果与处理日志导出。"""

import os
import re
import sys
import time
from copy import copy
from datetime import date

from openpyxl import Workbook, load_workbook

from logging_utils import print_log
from parsers import get_core_headers, get_order_type_value


def _resource_path(filename):
    """返回模板文件在源码目录或 PyInstaller 解包目录中的路径。"""
    if getattr(sys, "frozen", False):
        base_dir = getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_dir, filename)


def get_output_dir():
    """返回没有历史记录时目录选择框默认打开的导出目录。"""
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.getcwd()


EXPORT_DIR_STATE_FILE = os.path.expanduser("~/.ge_tool_export_dir")


def get_last_export_dir():
    """返回上次人工选择的导出目录，没有记录或不可读时返回 None。"""
    try:
        with open(EXPORT_DIR_STATE_FILE, encoding="utf-8") as state_file:
            export_dir = state_file.read().strip()
    except OSError:
        return None
    return export_dir or None


def save_last_export_dir(export_dir):
    """保存上次人工选择的导出目录，供下次启动作为默认位置。"""
    try:
        with open(EXPORT_DIR_STATE_FILE, "w", encoding="utf-8") as state_file:
            state_file.write(export_dir)
    except OSError:
        print_log("导出目录未能保存，下次启动不会记住")


PO_TEMPLATE_PATH = _resource_path("DOC_PO_HEADER.xlsx")
SALESORDER_TEMPLATE_PATH = _resource_path("DOC_SALESORDER_HEADER.xlsx")
OSCAR_SALESORDER_TEMPLATE_PATH = _resource_path("DOC_SALESORDER_HEADER_1.xlsx")

_DATE_PATTERN = re.compile(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$")

_INVOICE_FIXED_VALUES = {
    "A": "WH004078",
    "C": "00",
    "D": "GEHC",
    "R": "WH004078",
    "S": "GEHC",
    "V": "EA",
    "AC": "GOOD",
}

_INVOICE_DYNAMIC_COLUMNS = {
    "B": 0,   # 订单类型 -> OSI/POIN/REPAIRIN
    "G": 1,   # 运单号 -> poReference2
    "F": 2,   # INVOICE NO
    "L": 14,  # CARRIER
    "M": 15,  # HAWB
    "T": 3,   # ITEM NUMBER
    "U": 4,   # QTY
    "X": 8,   # Expiration Date
    "Z": 7,   # LOT Number
    "AD": 6,  # Serial Number
    "AF": 5,  # LPN Number
    "AJ": 9,  # COUNTRY OF ORIGIN
    "AK": 10, # SALES ORDER NO
    "AL": 11, # CUSTOMER PO
}

_SALESORDER_FIXED_VALUES = {
    "A": "WH004078",
    "B": "00",
    "E": "Y",
    "F": "GEHC",
    "V": "CONSIGNEEID",
    "AE": "WH004078",
    "AF": "GEHC",
    "AH": "00",
    "AJ": "ORACLE",  # 货物来源
    "AP": "EA",
    "AQ": "HD78_E841_01",
    "AR": "HD78_E841_01",
    "AS": "0",
    "AT": "0",
    "AU": "0",
    "AV": "0",
}

_SALESORDER_DYNAMIC_COLUMNS = {
    "C": 0,    # 订单类型 -> GNCK_*/GWCK_*
    "D": 25,   # Pick Slip Print Date
    "G": 1,    # Order Number
    "I": 26,   # System Id -> 参考编号3
    "L": 10,   # OrderType
    "M": 11,   # Ordered Date
    "N": 12,   # Shipment Priority
    "O": 13,   # Ship Method
    "P": 14,   # Service Level
    "Q": 15,   # FE SSO
    "R": 16,   # FE Name
    "S": 22,   # Shipping Instruction
    "T": 23,   # Special Instruction
    "U": 27,   # Pick From Subinv
    "W": 20,   # Ship To Address
    "Y": 19,   # SHIP TO NO -> udf01
    "AG": 3,   # Item Number
    "AI": 7,   # Lot
    "AK": 24,  # Org
    "AL": 27,  # Pick From Subinv -> 质量状态
    "AM": 6,   # Serial
    "AN": 5,   # LPN
    "AO": 4,   # Qty
    "AW": 2,   # Task Id
    "AY": 9,   # Pick From Locator
}

_OSCAR_FIXED_VALUES = {
    "A": "WH004078",
    "B": "00",
    "E": "Y",
    "F": "GEHC",
    "Z": "CONSIGNEEID",
    "AI": "WH004078",
    "AJ": "GEHC",
    "AL": "00",
    "AN": "OSCAR",  # 货物来源
    "AT": "EA",
    "AU": "HD78_E841_01",
    "AV": "HD78_E841_01",
    "AW": "0",
    "AX": "0",
    "AY": "0",
    "AZ": "0",
}

_OSCAR_DYNAMIC_COLUMNS = {
    "C": 0,    # 订单类型 -> GNCK_*/GWCK_*
    "G": 1,    # 服务申请号
    "H": 17,   # SR编号
    "I": 16,   # 客户设备id
    "P": 13,   # 时效
    "Q": 9,    # SSO
    "R": 11,   # 姓名
    "V": 8,    # 供应商
    "W": 10,   # 收货人
    "X": 14,   # 收货人电话
    "AA": 12,  # 收货地址
    "AB": 15,  # 申请说明
    "AK": 2,   # 物料编号
    "AO": 7,   # 仓库
    "AP": 6,   # 状态: 好件->GOOD
    "AQ": 4,   # 序列号
    "AS": 3,   # 数量
    "BB": 18,  # 跟踪号
    "BC": 5,   # 货位
}

_ORACLE_MONTH_NAMES = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}

_ORACLE_DATE_PATTERNS = (
    re.compile(
        r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})"
        r"(?:[ T](\d{1,2}):(\d{2})(?::(\d{2}))?)?$"
    ),
    re.compile(
        r"^(\d{1,2})-([A-Za-z]+)-(\d{2}|\d{4})"
        r"(?:[ T](\d{1,2}):(\d{2})(?::(\d{2}))?)?$"
    ),
)


def _row_value(row, index):
    """读取预览行字段，兼容界面手工新增的不完整行。"""
    return row[index] if index < len(row) else ""


def _normalize_expiration_date(value):
    """把 YYYY/MM/DD 或 YYYY-MM-DD 转成模板要求的 YYYY-MM-DD。"""
    if value is None:
        return ""
    text = str(value).strip()
    match = _DATE_PATTERN.fullmatch(text)
    if not match:
        return text
    try:
        year, month, day = (int(part) for part in match.groups())
        return date(year, month, day).strftime("%Y-%m-%d")
    except ValueError:
        return text


def _normalize_oracle_datetime(value, include_time):
    """把 Oracle 日期样例转换为销售订单模板要求的格式。"""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    for pattern in _ORACLE_DATE_PATTERNS:
        match = pattern.fullmatch(text)
        if not match:
            continue
        day_or_year = match.group(1)
        token2 = match.group(2)
        token3 = match.group(3)
        if pattern is _ORACLE_DATE_PATTERNS[0]:
            year, month, day = int(day_or_year), int(token2), int(token3)
        else:
            month_name = token2.upper()
            month_num = _ORACLE_MONTH_NAMES.get(month_name)
            if month_num is None:
                month_num = _ORACLE_MONTH_NAMES.get(month_name[:3])
            if month_num is None:
                return text
            day = int(day_or_year)
            year = int(token3)
            if len(token3) == 2:
                year = 2000 + year if year < 70 else 1900 + year
            month = month_num
        try:
            base = date(year, month, day).strftime("%Y-%m-%d")
        except ValueError:
            return text
        if match.group(4):
            time_text = f"{int(match.group(4)):02d}:{match.group(5)}"
            if match.group(6):
                time_text += f":{match.group(6)}"
            return f"{base} {time_text}" if include_time else base
        return f"{base} 00:00:00" if include_time else base
    return text


def _oracle_quality_status(pick_from_subinv):
    """按模板规则把 Pick From Subinv 后缀转换为质量状态。"""
    text = str(pick_from_subinv).strip().upper()
    if text.endswith("GD"):
        return "GOOD"
    if text.endswith("BAD"):
        return "BAD"
    return ""


def _normalize_oscar_serial(value):
    """把 OSCAR 序列号的空值/N/A 归一化为导出与发送空值。"""
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.upper() == "N/A":
        return ""
    return text


def _export_oracle_salesorder_template(core_data, output_file):
    """按 DOC_SALESORDER_HEADER 的销售订单表头模板导出售后单数据。"""
    wb = load_workbook(SALESORDER_TEMPLATE_PATH)
    ws = wb["0"]
    del wb["系统代码说明"]

    for row_index, preview_row in enumerate(core_data, start=3):
        if row_index > 3:
            for col_index in range(1, ws.max_column + 1):
                template_cell = ws.cell(row=3, column=col_index)
                target_cell = ws.cell(row=row_index, column=col_index)
                target_cell._style = copy(template_cell._style)

        for col_name, value in _SALESORDER_FIXED_VALUES.items():
            ws[f"{col_name}{row_index}"] = value
        for col_name, source_index in _SALESORDER_DYNAMIC_COLUMNS.items():
            value = _row_value(preview_row, source_index)
            if col_name == "C":
                value = get_order_type_value("GE-ORACLE拣货单", value)
            elif col_name == "D":
                value = _normalize_oracle_datetime(value, include_time=True)
            elif col_name == "M":
                value = _normalize_oracle_datetime(value, include_time=False)
            elif col_name == "AL":
                value = _oracle_quality_status(value)
            ws[f"{col_name}{row_index}"] = value

    wb.save(output_file)


def _export_oscar_salesorder_template(core_data, output_file):
    """按 DOC_SALESORDER_HEADER_1 的销售订单表头模板导出OSCAR拣货单数据。"""
    wb = load_workbook(OSCAR_SALESORDER_TEMPLATE_PATH)
    ws = wb["0"]
    del wb["系统代码说明"]

    creation_time = time.strftime("%Y-%m-%d %H:%M:%S")

    for row_index, preview_row in enumerate(core_data, start=3):
        if row_index > 3:
            for col_index in range(1, ws.max_column + 1):
                template_cell = ws.cell(row=3, column=col_index)
                target_cell = ws.cell(row=row_index, column=col_index)
                target_cell._style = copy(template_cell._style)

        ws[f"D{row_index}"] = creation_time
        for col_name, value in _OSCAR_FIXED_VALUES.items():
            ws[f"{col_name}{row_index}"] = value
        for col_name, source_index in _OSCAR_DYNAMIC_COLUMNS.items():
            value = _row_value(preview_row, source_index)
            if col_name == "C":
                value = get_order_type_value("GE-OSCAR拣货单", value)
            elif col_name == "AP":
                value = "GOOD" if str(value).strip() == "好件" else ""
            elif col_name == "AQ":
                value = _normalize_oscar_serial(value)
            ws[f"{col_name}{row_index}"] = value

    wb.save(output_file)


def _export_invoice_po_template(core_data, output_file):
    """按 DOC_PO_HEADER 的采购订单表头模板导出发票数据。"""
    wb = load_workbook(PO_TEMPLATE_PATH)
    ws = wb["采购订单表头"]
    del wb["系统代码说明"]

    now = time.localtime()
    creation_time = f"{now.tm_year}-{now.tm_mon}-{now.tm_mday} " \
                    f"{time.strftime('%H:%M:%S', now)}"

    for row_index, preview_row in enumerate(core_data, start=3):
        if row_index > 3:
            for col_index in range(1, ws.max_column + 1):
                template_cell = ws.cell(row=3, column=col_index)
                target_cell = ws.cell(row=row_index, column=col_index)
                target_cell._style = copy(template_cell._style)

        ws[f"E{row_index}"] = creation_time
        for col_name, value in _INVOICE_FIXED_VALUES.items():
            ws[f"{col_name}{row_index}"] = value
        for col_name, source_index in _INVOICE_DYNAMIC_COLUMNS.items():
            value = _row_value(preview_row, source_index)
            if col_name == "X":
                value = _normalize_expiration_date(value)
            elif col_name == "B":
                value = get_order_type_value("GE-发票单", value)
            ws[f"{col_name}{row_index}"] = value
        if ws[f"B{row_index}"].value == "OSI":
            ws[f"AA{row_index}"] = "ORACLE"

    wb.save(output_file)


def export_excel(select_text, core_data, full_log_data, output_file=None):
    """生成单个文件的识别结果 Excel，返回输出文件路径。"""
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    output_file = output_file or os.path.join(
        get_output_dir(), f"GE单据OCR识别结果_{timestamp}.xlsx"
    )

    if select_text == "GE-发票单":
        _export_invoice_po_template(core_data, output_file)
        print_log(f"Excel导出完成，路径：{output_file}")
        return output_file
    if select_text == "GE-ORACLE拣货单":
        _export_oracle_salesorder_template(core_data, output_file)
        print_log(f"Excel导出完成，路径：{output_file}")
        return output_file
    if select_text == "GE-OSCAR拣货单":
        _export_oscar_salesorder_template(core_data, output_file)
        print_log(f"Excel导出完成，路径：{output_file}")
        return output_file

    wb = Workbook()
    ws_core = wb.active
    ws_core.title = "识别结果列表"

    core_headers = get_core_headers(select_text)

    ws_core.append(core_headers)
    for row in core_data:
        ws_core.append(row)

    # 自适应列宽
    for col in ws_core.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        ws_core.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    # 日志Sheet
    ws_log = wb.create_sheet(title="处理日志")
    log_headers = ["文件名", "fileId", "文件URL", "reqUuid",
                   "提取单据编号", "状态", "异常信息", "识别报文"]
    ws_log.append(log_headers)
    for row in full_log_data:
        ws_log.append(row)
    for col in ws_log.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        ws_log.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)

    wb.save(output_file)
    print_log(f"Excel导出完成，路径：{output_file}")
    return output_file
